'''
Name: trait_extract_parallel.py

Version: 1.0

Summary: Extract plant shoot traits (larea, solidity, max_width, max_height, avg_curv, color_cluster) by paralell processing 
    
Author: suxing liu

Author-email: suxingliu@gmail.com

Created: 2018-09-29

USAGE:

time python3 trait_extract_parallel.py -p ~/plant-image-analysis/test/ -ft jpg 

time python3 trait_extract_parallel.py -p ~/plant-image-analysis/test/ -ft jpg -r /home/suxingliu/plant-image-analysis/test/results/

'''

import argparse
import csv
import glob
# import the necessary packages
import os
import traceback
import warnings
from collections import Counter
from os.path import join
from typing import List

import cv2
import imutils
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from scipy import ndimage
from scipy import optimize
from scipy.interpolate import interp1d
from scipy.spatial import distance as dist
from skan import Skeleton, summarize, draw
from skimage import img_as_float, img_as_ubyte, img_as_bool
from skimage import morphology
from skimage.color import rgb2lab, deltaE_cie76
from skimage.feature import peak_local_max
from skimage.segmentation import clear_border, watershed
from sklearn.cluster import KMeans
from tabulate import tabulate

from core.luminous_detection import isbright, write_results_to_csv
from core.options import ImageInput
from core.results import ImageResult
from core.thresholding import otsu_threshold

warnings.filterwarnings("ignore")

from pathlib import Path

MBFACTOR = float(1<<20)

class ComputeCurvature:

    def __init__(self,x,y):
        """ Initialize some variables """
        self.xc = 0  # X-coordinate of circle center
        self.yc = 0  # Y-coordinate of circle center
        self.r = 0   # Radius of the circle
        self.xx = np.array([])  # Data points
        self.yy = np.array([])  # Data points
        self.x = x  # X-coordinate of circle center
        self.y = y  # Y-coordinate of circle center

    def calc_r(self, xc, yc):
        """ calculate the distance of each 2D points from the center (xc, yc) """
        return np.sqrt((self.xx-xc)**2 + (self.yy-yc)**2)

    def f(self, c):
        """ calculate the algebraic distance between the data points and the mean circle centered at c=(xc, yc) """
        ri = self.calc_r(*c)
        return ri - ri.mean()

    def df(self, c):
        """ Jacobian of f_2b
        The axis corresponding to derivatives must be coherent with the col_deriv option of leastsq"""
        xc, yc = c
        df_dc = np.empty((len(c), self.x.size))

        ri = self.calc_r(xc, yc)
        df_dc[0] = (xc - self.x)/ri                   # dR/dxc
        df_dc[1] = (yc - self.y)/ri                   # dR/dyc
        df_dc = df_dc - df_dc.mean(axis=1)[:, np.newaxis]
        return df_dc

    def fit(self, xx, yy):
        self.xx = xx
        self.yy = yy
        center_estimate = np.r_[np.mean(xx), np.mean(yy)]
        center = optimize.leastsq(self.f, center_estimate, Dfun=self.df, col_deriv=True)[0]

        self.xc, self.yc = center
        ri = self.calc_r(*center)
        self.r = ri.mean()

        return 1 / self.r  # Return the curvature


# generate foloder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        return False
        

def color_cluster_seg(image, args_colorspace, args_channels, args_num_clusters):
    
    # Change image color space, if necessary.
    colorSpace = args_colorspace.lower()

    if colorSpace == 'hsv':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    elif colorSpace == 'ycrcb' or colorSpace == 'ycc':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2YCrCb)

    elif colorSpace == 'lab':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)

    else:
        colorSpace = 'bgr'  # set for file naming purposes
        image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

    # Keep only the selected channels for K-means clustering.
    if args_channels != 'all':
        channels = cv2.split(image)
        channelIndices = []
        for char in args_channels:
            channelIndices.append(int(char))
        image = image[:,:,channelIndices]
        if len(image.shape) == 2:
            image.reshape(image.shape[0], image.shape[1], 1)
            
    (width, height, n_channel) = image.shape
    
    #print("image shape: \n")
    #print(width, height, n_channel)
    
 
    # Flatten the 2D image array into an MxN feature vector, where M is the number of pixels and N is the dimension (number of channels).
    reshaped = image.reshape(image.shape[0] * image.shape[1], image.shape[2])
    

    # Perform K-means clustering.
    if args_num_clusters < 2:
        print('Warning: num-clusters < 2 invalid. Using num-clusters = 2')
    
    #define number of cluster
    numClusters = max(2, args_num_clusters)

    # clustering method
    kmeans = KMeans(n_clusters = numClusters, n_init = 40, max_iter = 500).fit(reshaped)
    
    # get lables 
    pred_label = kmeans.labels_
    
    # Reshape result back into a 2D array, where each element represents the corresponding pixel's cluster index (0 to K - 1).
    clustering = np.reshape(np.array(pred_label, dtype=np.uint8), (image.shape[0], image.shape[1]))

    # Sort the cluster labels in order of the frequency with which they occur.
    sortedLabels = sorted([n for n in range(numClusters)],key = lambda x: -np.sum(clustering == x))

    # Initialize K-means grayscale image; set pixel colors based on clustering.
    kmeansImage = np.zeros(image.shape[:2], dtype=np.uint8)
    for i, label in enumerate(sortedLabels):
        kmeansImage[clustering == label] = int(255 / (numClusters - 1)) * i

    thresh = otsu_threshold(kmeansImage)

    # dilate the binary image
    kernel = np.ones((5, 5), np.uint8)
    # thresh = cv2.dilate(thresh, kernel, iterations=1)
    #cv2.imwrite(join(options.output_directory, f"{options.input_stem}_thresh{file_extension}"), thresholded)
    
    # ret, thresh = cv2.threshold(kmeansImage, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    #return thresh =
    
    #num_of_non_zeros = np.count_nonzero(thresh)
    
    thresh_cleaned = clear_border(thresh)
    
    if np.count_nonzero(thresh) > 0:
        
        thresh_cleaned_bw = clear_border(thresh)
    else:
        thresh_cleaned_bw = thresh
        
    nb_components, output, stats, centroids = cv2.connectedComponentsWithStats(thresh_cleaned_bw, connectivity = 8)
    
    # stats[0], centroids[0] are for the background label. ignore
    # cv2.CC_STAT_LEFT, cv2.CC_STAT_TOP, cv2.CC_STAT_WIDTH, cv2.CC_STAT_HEIGHT
    sizes = stats[1:, cv2.CC_STAT_AREA]
    
    Coord_left = stats[1:, cv2.CC_STAT_LEFT]
    
    Coord_top = stats[1:, cv2.CC_STAT_TOP]
    
    Coord_width = stats[1:, cv2.CC_STAT_WIDTH]
    
    Coord_height = stats[1:, cv2.CC_STAT_HEIGHT]
    
    Coord_centroids = centroids
    
    #print("Coord_centroids {}\n".format(centroids[1][1]))
    
    #print("[width, height] {} {}\n".format(width, height))
    
    nb_components = nb_components - 1
    
    min_size = 1000 
    
    max_size = width*height*0.1
    
    img_thresh = np.zeros([width, height], dtype=np.uint8)
    
    
    for i in range(0, nb_components):
        
        if (sizes[i] >= min_size):
            
            if (Coord_left[i] > 1) and (Coord_top[i] > 1) and (Coord_width[i] - Coord_left[i] > 0) and (Coord_height[i] - Coord_top[i] > 0) and (centroids[i][0] - width*0.5 < 10) and (centroids[i][1] - height*0.5 < 10):
                img_thresh[output == i + 1] = 255
                print("Foreground center found ")
            
            elif ((Coord_width[i] - Coord_left[i])*0.5 - width < 15) and (centroids[i][0] - width*0.5 < 15) and (centroids[i][1] - height*0.5 < 15) and ((sizes[i] <= max_size)):
                imax = max(enumerate(sizes), key=(lambda x: x[1]))[0] + 1    
                img_thresh[output == imax] = 255
                print("Foreground max found ")
            
            else:
                img_thresh[output == i + 1] = 255
    
    #from skimage import img_as_ubyte
    
    #img_thresh = img_as_ubyte(img_thresh)
    
    #print("img_thresh.dtype")
    #print(img_thresh.dtype)
    
    return img_thresh
    
    #return thresh_cleaned
    
'''
def medial_axis_image(thresh):
    
    #convert an image from OpenCV to skimage
    thresh_sk = img_as_float(thresh)

    image_bw = img_as_bool((thresh_sk))
    
    image_medial_axis = medial_axis(image_bw)
    
    return image_medial_axis
'''

def skeleton_bw(thresh):

    # Convert mask to boolean image, rather than 0 and 255 for skimage to use it
    
    #convert an image from OpenCV to skimage
    thresh_sk = img_as_float(thresh)

    image_bw = img_as_bool((thresh_sk))

    skeleton = morphology.skeletonize(image_bw)

    skeleton_img = skeleton.astype(np.uint8) * 255

    return skeleton_img, skeleton


def watershed_seg(orig, thresh, min_distance_value):
    
    # compute the exact Euclidean distance from every binary
    # pixel to the nearest zero pixel, then find peaks in this
    # distance map
    D = ndimage.distance_transform_edt(thresh)
    
    localMax = peak_local_max(D, indices = False, min_distance = min_distance_value,  labels = thresh)
     
    # perform a connected component analysis on the local peaks,
    # using 8-connectivity, then appy the Watershed algorithm
    markers = ndimage.label(localMax, structure = np.ones((3, 3)))[0]
    
    #print("markers")
    #print(type(markers))
    
    labels = watershed(-D, markers, mask = thresh)
    
    print("[INFO] {} unique segments found\n".format(len(np.unique(labels)) - 1))
    
    return labels


def individual_object_seg(orig, labels, save_path, base_name, file_extension, leaf_images: bool = True):
    (width, height, n_channel) = orig.shape
    
    for label in np.unique(labels):
    # if the label is zero, we are examining the 'background'
    # so simply ignore it
        if label == 0:
            continue
     
        # otherwise, allocate memory for the label region and draw
        # it on the mask
        mask = np.zeros((width, height), dtype="uint8")
        mask[labels == label] = 255
        
        # apply individual object mask
        masked = cv2.bitwise_and(orig, orig, mask = mask)
        
        if leaf_images:
            result_img_path = (save_path + base_name + '_leaf_' + str(label) + file_extension)
            cv2.imwrite(result_img_path, masked)
        


'''
def watershed_seg_marker(orig, thresh, min_distance_value, img_marker):
    
    # compute the exact Euclidean distance from every binary
    # pixel to the nearest zero pixel, then find peaks in this
    # distance map
    D = ndimage.distance_transform_edt(thresh)
    
    gray = cv2.cvtColor(img_marker, cv2.COLOR_BGR2GRAY)
    img_marker = cv2.threshold(gray, 128, 255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    #localMax = peak_local_max(D, indices = False, min_distance = min_distance_value,  labels = thresh)
     
    # perform a connected component analysis on the local peaks,
    # using 8-connectivity, then appy the Watershed algorithm
    markers = ndimage.label(img_marker, structure = np.ones((3, 3)))[0]

    labels = watershed(-D, markers, mask = thresh)

    
    props = regionprops(labels)
    
    areas = [p.area for p in props]
    
    import statistics
    
    
    #outlier_list = outlier_doubleMAD(areas, thresh = 1.0)
    
    #indices = [i for i, x in enumerate(outlier_list) if x]
    
    print(areas)
    print(statistics.mean(areas))
    #
    #print(outlier_list)
    #print(indices)
    
    
    print("[INFO] {} unique segments found\n".format(len(np.unique(labels)) - 1))
    
    return labels
'''


def comp_external_contour(orig,thresh):
    
    #find contours and get the external one
    contours, hier = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
   
    img_height, img_width, img_channels = orig.shape
   
    index = 1
    
    for c in contours:
        
        #get the bounding rect
        x, y, w, h = cv2.boundingRect(c)
        
        if w>img_width*0.01 and h>img_height*0.01:
            
            trait_img = cv2.drawContours(orig, contours, -1, (255, 255, 0), 1)
    
            # draw a green rectangle to visualize the bounding rect
            roi = orig[y:y+h, x:x+w]
            
            print("ROI {} detected ...\n".format(index))
            #result_file = (save_path +  str(index) + file_extension)
            #cv2.imwrite(result_file, roi)
            
            trait_img = cv2.rectangle(orig, (x, y), (x+w, y+h), (255, 255, 0), 3)
            
            index+= 1

            '''
            #get the min area rect
            rect = cv2.minAreaRect(c)
            box = cv2.boxPoints(rect)
            # convert all coordinates floating point values to int
            box = np.int0(box)
            #draw a red 'nghien' rectangle
            trait_img = cv2.drawContours(orig, [box], 0, (0, 0, 255))
            '''
             # get convex hull
            hull = cv2.convexHull(c)
            # draw it in red color
            trait_img = cv2.drawContours(orig, [hull], -1, (0, 0, 255), 3)
            
            '''
            # calculate epsilon base on contour's perimeter
            # contour's perimeter is returned by cv2.arcLength
            epsilon = 0.01 * cv2.arcLength(c, True)
            # get approx polygons
            approx = cv2.approxPolyDP(c, epsilon, True)
            # draw approx polygons
            trait_img = cv2.drawContours(orig, [approx], -1, (0, 255, 0), 1)
         
            # hull is convex shape as a polygon
            hull = cv2.convexHull(c)
            trait_img = cv2.drawContours(orig, [hull], -1, (0, 0, 255))
            '''
            
            '''
            #get the min enclosing circle
            (x, y), radius = cv2.minEnclosingCircle(c)
            # convert all values to int
            center = (int(x), int(y))
            radius = int(radius)
            # and draw the circle in blue
            trait_img = cv2.circle(orig, center, radius, (255, 0, 0), 2)
            '''
            
            area = cv2.contourArea(c)
            print("Leaf area = {0:.2f}... \n".format(area))
            
            
            hull = cv2.convexHull(c)
            hull_area = cv2.contourArea(hull)
            solidity = float(area)/hull_area
            print("solidity = {0:.2f}... \n".format(solidity))
            
            extLeft = tuple(c[c[:,:,0].argmin()][0])
            extRight = tuple(c[c[:,:,0].argmax()][0])
            extTop = tuple(c[c[:,:,1].argmin()][0])
            extBot = tuple(c[c[:,:,1].argmax()][0])
            
            trait_img = cv2.circle(orig, extLeft, 3, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, extRight, 3, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, extTop, 3, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, extBot, 3, (255, 0, 0), -1)
            
            max_width = dist.euclidean(extLeft, extRight)
            max_height = dist.euclidean(extTop, extBot)
            
            if max_width > max_height:
                trait_img = cv2.line(orig, extLeft, extRight, (0,255,0), 2)
            else:
                trait_img = cv2.line(orig, extTop, extBot, (0,255,0), 2)

            print("Width and height are {0:.2f},{1:.2f}... \n".format(w, h))
            
            
            
    return trait_img, area, solidity, w, h
    
    

def compute_curv(orig, labels):
    
    gray = cv2.cvtColor(orig, cv2.COLOR_BGR2GRAY)
    
    curv_sum = 0.0
    count = 0
    # curvature computation
    # loop over the unique labels returned by the Watershed algorithm
    for index, label in enumerate(np.unique(labels), start = 1):
        # if the label is zero, we are examining the 'background'
        # so simply ignore it
        if label == 0:
            continue
     
        # otherwise, allocate memory for the label region and draw
        # it on the mask
        mask = np.zeros(gray.shape, dtype = "uint8")
        mask[labels == label] = 255
     
        # detect contours in the mask and grab the largest one
        #cnts = cv2.findContours(mask.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[-2]
        contours, hierarchy = cv2.findContours(mask.copy(),cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        c = max(contours, key = cv2.contourArea)
        
        # draw a circle enclosing the object
        ((x, y), r) = cv2.minEnclosingCircle(c)
        label_trait = cv2.circle(orig, (int(x), int(y)), 3, (0, 255, 0), 2)
        label_trait = cv2.putText(orig, "#{}".format(label), (int(x) - 10, int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
        #cv2.putText(orig, "#{}".format(curvature), (int(x) - 10, int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
        
        
        if len(c) >= 5 :
            try:
                label_trait = cv2.drawContours(orig, [c], -1, (255, 0, 0), 2)
                ellipse = cv2.fitEllipse(c)
                label_trait = cv2.ellipse(orig,ellipse,(0,255,0),2)

                c_np = np.vstack(c).squeeze()
                count+=1

                x = c_np[:,0]
                y = c_np[:,1]

                comp_curv = ComputeCurvature(x, y)
                curvature = comp_curv.fit(x, y)

                curv_sum = curv_sum + curvature
            except:
                print(traceback.format_exc())
        else:
            # optional to "delete" the small contours
            label_trait = cv2.drawContours(orig, [c], -1, (0, 0, 255), 2)
            print("lack of enough points to fit ellipse")
    
    if count > 0:
        print('average curvature = {0:.2f}\n'.format(curv_sum/count))
    else:
        count = 1.0
    
    return curv_sum/count, label_trait


def RGB2HEX(color):
    return "#{:02x}{:02x}{:02x}".format(int(color[0]), int(color[1]), int(color[2]))

'''
def color_quantization(image, mask, save_path, num_clusters):
    
    #grab image width and height
    (h, w) = image.shape[:2]
    
    #change the color storage order
    image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    
    #apply the mask to get the segmentation of plant
    masked_image = cv2.bitwise_and(image, image, mask = mask)
       
    # reshape the image to be a list of pixels
    pixels = masked_image.reshape((masked_image.shape[0] * masked_image.shape[1], 3))
        
    ############################################################
    #Clustering process
    ###############################################################
    # cluster the pixel intensities
    clt = MiniBatchKMeans(n_clusters = num_clusters)
    #clt = KMeans(n_clusters = args["clusters"])
    clt.fit(pixels)

    #assign labels to each cluster 
    labels = clt.fit_predict(pixels)

    #obtain the quantized clusters using each label
    quant = clt.cluster_centers_.astype("uint8")[labels]

    # reshape the feature vectors to images
    quant = quant.reshape((h, w, 3))
    image_rec = pixels.reshape((h, w, 3))
    
    # convert from L*a*b* to RGB
    quant = cv2.cvtColor(quant, cv2.COLOR_RGB2BGR)
    image_rec = cv2.cvtColor(image_rec, cv2.COLOR_RGB2BGR)
    
    # display the images 
    #cv2.imshow("image", np.hstack([image_rec, quant]))
    #cv2.waitKey(0)
    
    #define result path for labeled images
    result_img_path = save_path + 'cluster_out.png'
    
    # save color_quantization results
    cv2.imwrite(result_img_path,quant)

    #Get colors and analze them from masked image
    counts = Counter(labels)
    # sort to ensure correct color percentage
    counts = dict(sorted(counts.items()))
    
    center_colors = clt.cluster_centers_
    
    #print(type(center_colors))

    # We get ordered colors by iterating through the keys
    ordered_colors = [center_colors[i] for i in counts.keys()]
    hex_colors = [RGB2HEX(ordered_colors[i]) for i in counts.keys()]
    rgb_colors = [ordered_colors[i] for i in counts.keys()]
    
    
    #######################################################################################
    threshold = 60
    
    selected_color = rgb2lab(np.uint8(np.asarray([[rgb_colors[0]]])))

    for i in range(num_clusters):
        curr_color = rgb2lab(np.uint8(np.asarray([[rgb_colors[i]]]))) 
        diff = deltaE_cie76(selected_color, curr_color)
        if (diff < threshold):
            print("Color difference value is : {0} \n".format(str(diff)))
    ###########################################################################################
    #print(hex_colors)
    
    index_bkg = [index for index in range(len(hex_colors)) if hex_colors[index] == '#000000']
    
    #print(index_bkg[0])

    #print(counts)
    #remove background color 
    del hex_colors[index_bkg[0]]
    del rgb_colors[index_bkg[0]]
    
    # Using dictionary comprehension to find list 
    # keys having value . 
    delete = [key for key in counts if key == index_bkg[0]] 
  
    # delete the key 
    for key in delete: del counts[key] 
   
    fig = plt.figure(figsize = (6, 6))
    plt.pie(counts.values(), labels = hex_colors, colors = hex_colors)

    #define result path for labeled images
    result_img_path = save_path + 'pie_color.png'
    plt.savefig(result_img_path)
        
    # build a histogram of clusters and then create a figure representing the number of pixels labeled to each color
    hist = utils.centroid_histogram(clt)

    # remove the background color cluster
    clt.cluster_centers_ = np.delete(clt.cluster_centers_, index_bkg[0], axis=0)
    
    #build a histogram of clusters using center lables
    numLabels = utils.plot_centroid_histogram(save_path,clt)

    #create a figure representing the distribution of each color
    bar = utils.plot_colors(hist, clt.cluster_centers_)

    #save a figure of color bar 
    utils.plot_color_bar(save_path, bar)

    return rgb_colors
'''

def get_cmap(n, name='hsv'):
    '''Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    RGB color; the keyword argument name must be a standard mpl colormap name.'''
    return plt.cm.get_cmap(name, n)
    
    

def color_region(image, mask, output_directory, file_name, num_clusters):
    
    # read the image
     #grab image width and height
    (h, w) = image.shape[:2]

    #apply the mask to get the segmentation of plant
    masked_image_ori = cv2.bitwise_and(image, image, mask = mask)
    
    #define result path for labeled images
    result_img_path = join(output_directory, f"{file_name}.masked.png")
    cv2.imwrite(result_img_path, masked_image_ori)
    
    # convert to RGB
    image_RGB = cv2.cvtColor(masked_image_ori, cv2.COLOR_BGR2RGB)

    # reshape the image to a 2D array of pixels and 3 color values (RGB)
    pixel_values = image_RGB.reshape((-1, 3))
    
    # convert to float
    pixel_values = np.float32(pixel_values)

    # define stopping criteria
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.2)

    # number of clusters (K)
    #num_clusters = 5
    compactness, labels, (centers) = cv2.kmeans(pixel_values, num_clusters, None, criteria, 10, cv2.KMEANS_RANDOM_CENTERS)

    # convert back to 8 bit values
    centers = np.uint8(centers)

    # flatten the labels array
    labels_flat = labels.flatten()

    # convert all pixels to the color of the centroids
    segmented_image = centers[labels_flat]

    # reshape back to the original image dimension
    segmented_image = segmented_image.reshape(image_RGB.shape)


    segmented_image_BRG = cv2.cvtColor(segmented_image, cv2.COLOR_RGB2BGR)
    #define result path for labeled images
    result_img_path = join(output_directory, f"{file_name}.clustered.png")
    cv2.imwrite(result_img_path, segmented_image_BRG)


    '''
    fig = plt.figure()
    ax = Axes3D(fig)        
    for label, pix in zip(labels, segmented_image):
        ax.scatter(pix[0], pix[1], pix[2], color = (centers))
            
    result_file = (save_path + base_name + 'color_cluster_distributation.png')
    plt.savefig(result_file)
    '''
    #Show only one chosen cluster 
    #masked_image = np.copy(image)
    masked_image = np.zeros_like(image_RGB)

    # convert to the shape of a vector of pixel values
    masked_image = masked_image.reshape((-1, 3))
    # color (i.e cluster) to render
    #cluster = 2

    cmap = get_cmap(num_clusters+1)
    
    #clrs = sns.color_palette('husl', n_colors = num_clusters)  # a list of RGB tuples

    color_conversion = interp1d([0,1],[0,255])


    for cluster in range(num_clusters):

        print("Processing Cluster{0} ...\n".format(cluster))
        #print(clrs[cluster])
        #print(color_conversion(clrs[cluster]))

        masked_image[labels_flat == cluster] = centers[cluster]

        #print(centers[cluster])

        #convert back to original shape
        masked_image_rp = masked_image.reshape(image_RGB.shape)

        #masked_image_BRG = cv2.cvtColor(masked_image, cv2.COLOR_RGB2BGR)
        #cv2.imwrite('maksed.png', masked_image_BRG)

        gray = cv2.cvtColor(masked_image_rp, cv2.COLOR_BGR2GRAY)

        # threshold the image, then perform a series of erosions +
        # dilations to remove any small regions of noise
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY)[1]

        #thresh_cleaned = clear_border(thresh)

        cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = imutils.grab_contours(cnts)
        #c = max(cnts, key=cv2.contourArea)

        '''
        # compute the center of the contour area and draw a circle representing the center
        M = cv2.moments(c)
        cX = int(M["m10"] / M["m00"])
        cY = int(M["m01"] / M["m00"])
        # draw the countour number on the image
        result = cv2.putText(masked_image_rp, "#{}".format(cluster + 1), (cX - 20, cY), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 255), 2)
        '''
        
        
        if not cnts:
            print("findContours is empty")
        else:
            
            # loop over the (unsorted) contours and draw them
            for (i, c) in enumerate(cnts):

                result = cv2.drawContours(masked_image_rp, c, -1, color_conversion(np.random.random(3)), 2)
                #result = cv2.drawContours(masked_image_rp, c, -1, color_conversion(clrs[cluster]), 2)

            #result = result(np.where(result == 0)== 255)
            result[result == 0] = 255


            result_BRG = cv2.cvtColor(result, cv2.COLOR_RGB2BGR)
            result_img_path = join(output_directory, f"{file_name}.result.{cluster}.png")
            cv2.imwrite(result_img_path, result_BRG)


    
    counts = Counter(labels_flat)
    # sort to ensure correct color percentage
    counts = dict(sorted(counts.items()))
    
    center_colors = centers

    # We get ordered colors by iterating through the keys
    ordered_colors = [center_colors[i] for i in counts.keys()]
    hex_colors = [RGB2HEX(ordered_colors[i]) for i in counts.keys()]
    rgb_colors = [ordered_colors[i] for i in counts.keys()]

    #print(hex_colors)
    
    index_bkg = [index for index in range(len(hex_colors)) if hex_colors[index] == '#000000']
    
    #print(index_bkg[0])

    #print(counts)
    #remove background color 
    del hex_colors[index_bkg[0]]
    del rgb_colors[index_bkg[0]]
    
    # Using dictionary comprehension to find list 
    # keys having value . 
    delete = [key for key in counts if key == index_bkg[0]] 
  
    # delete the key 
    for key in delete: del counts[key] 
   
    fig = plt.figure(figsize = (6, 6))
    plt.pie(counts.values(), labels = hex_colors, colors = hex_colors)

    #define result path for labeled images
    result_img_path = join(output_directory, f"{file_name}.pie_color.png")
    plt.savefig(result_img_path)

   
    return rgb_colors


def _normalise_image(image, *, image_cmap=None):
    image = img_as_float(image)
    if image.ndim == 2:
        if image_cmap is None:
            image = gray2rgb(image)
        else:
            image = plt.get_cmap(image_cmap)(image)[..., :3]
    return image


'''
def overlay_skeleton_endpoints(image, stats, *, image_cmap=None, axes=None):

    image = _normalise_image(image, image_cmap=image_cmap)
    summary = stats
    # transforming from row, col to x, y
    #coords_cols = (['image-coord-src-%i' % i for i in [1, 0]] +
    #               ['image-coord-dst-%i' % i for i in [1, 0]])
    
    coords_cols_src = (['image-coord-src-%i' % i for i in [1, 0]])
    coords_cols_dst = (['image-coord-dst-%i' % i for i in [1, 0]])
    
    #coords = summary[coords_cols].values.reshape((-1, 1, 2))
    
    coords_src = summary[coords_cols_src].values
    coords_dst = summary[coords_cols_dst].values

    coords_src_x = [i[0] for i in coords_src]
    coords_src_y = [i[1] for i in coords_src]
    
    coords_dst_x = [i[0] for i in coords_dst]
    coords_dst_y = [i[1] for i in coords_dst]
    
    img_marker = np.zeros_like(image, dtype = np.uint8)
    img_marker.fill(0) # or img[:] = 255
    img_marker[list(map(int, coords_src_y)), list(map(int, coords_src_x))] = 255
    img_marker[list(map(int, coords_dst_y)), list(map(int, coords_dst_x))] = 255
    
    #print("img_marker")
    #print(img_marker.shape)
    
    if axes is None:
        fig, axes = plt.subplots()
    
    axes.axis('off')
    axes.imshow(image)

    axes.scatter(coords_src_x, coords_src_y, c = 'w')
    axes.scatter(coords_dst_x, coords_dst_y, c = 'w')

    return fig, img_marker
    #return coords
'''

def outlier_doubleMAD(data,thresh = 3.5):
    
    """
    FOR ASSYMMETRIC DISTRIBUTION
    Returns : filtered array excluding the outliers

    Parameters : the actual data Points array

    Calculates median to divide data into 2 halves.(skew conditions handled)
    Then those two halves are treated as separate data with calculation same as for symmetric distribution.(first answer) 
    Only difference being , the thresholds are now the median distance of the right and left median with the actual data median
    """
    
    # warning: this function does not check for NAs
    # nor does it address issues when 
    # more than 50% of your data have identical values
    m = np.median(data)
    abs_dev = np.abs(data - m)
    left_mad = np.median(abs_dev[data <= m])
    right_mad = np.median(abs_dev[data >= m])
    data_mad = left_mad * np.ones(len(data))
    data_mad[data > m] = right_mad
    modified_z_score = 0.6745 * abs_dev / data_mad
    modified_z_score[data == m] = 0
    return modified_z_score > thresh




def extract_traits(image_file):
    
    #gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    abs_path = os.path.abspath(image_file)
    
    filename, file_extension = os.path.splitext(abs_path)
    base_name = os.path.splitext(os.path.basename(filename))[0]

    file_size = os.path.getsize(image_file)/MBFACTOR
    
    image_file_name = Path(image_file).name
   
    
    # make the folder to store the results
    #current_path = abs_path + '/'
    
    print("Exacting traits for image : {0}\n".format(str(image_file_name)))
     
    # save folder construction
    if (args['result']):
        save_path = args['result']
    else:
        mkpath = os.path.dirname(abs_path) +'/' + base_name
        mkdir(mkpath)
        save_path = mkpath + '/'

    print ("results_folder: " + save_path)

    print("Segmenting plant object using automatic color clustering method")
    if (file_size > 5.0):
        print(f"It may take some time due to large file size ({file_size} MB)")

    image = cv2.imread(image_file)
    
    #make backup image
    orig = image.copy()
    
    source_image = cv2.cvtColor(orig, cv2.COLOR_BGR2RGB)
     
    args_colorspace = args['color_space']
    args_channels = args['channels']
    args_num_clusters = args['num_clusters']
    
    #color clustering based plant object segmentation
    thresh = color_cluster_seg(orig, args_colorspace, args_channels, args_num_clusters)
    
    # save segmentation result
    result_file = (save_path + base_name + '_seg' + file_extension)
    #print(filename)
    cv2.imwrite(result_file, thresh)
    
    
    
    num_clusters = 5
    #save color quantization result
    #rgb_colors = color_quantization(image, thresh, save_path, num_clusters)
    rgb_colors = color_region(orig, thresh, save_path, filename, num_clusters)
    
    selected_color = rgb2lab(np.uint8(np.asarray([[rgb_colors[0]]])))
    
    print("Color difference are : ") 
    
    print(selected_color)
    
    color_diff = []
    
    for index, value in enumerate(rgb_colors): 
        #print(index, value) 
        curr_color = rgb2lab(np.uint8(np.asarray([[value]])))
        diff = deltaE_cie76(selected_color, curr_color)
        
        color_diff.append(diff)
        
        print(index, value, diff) 
    
    
    
    ###############################################
    
    #accquire medial axis of segmentation mask
    #image_medial_axis = medial_axis_image(thresh)
    
    image_skeleton, skeleton = skeleton_bw(thresh)

    # save _skeleton result
    result_file = (save_path + base_name + '_skeleton' + file_extension)
    cv2.imwrite(result_file, img_as_ubyte(image_skeleton))

    ###
    # ['skeleton-id', 'node-id-src', 'node-id-dst', 'branch-distance', 
    #'branch-type', 'mean-pixel-value', 'stdev-pixel-value', 
    #'image-coord-src-0', 'image-coord-src-1', 'image-coord-dst-0', 'image-coord-dst-1', 
    #'coord-src-0', 'coord-src-1', 'coord-dst-0', 'coord-dst-1', 'euclidean-distance']
    ###
    
    #get brach data
    branch_data = summarize(Skeleton(image_skeleton))
    #print(branch_data)
    
    #select end branch
    sub_branch = branch_data.loc[branch_data['branch-type'] == 1]
    
    sub_branch_branch_distance = sub_branch["branch-distance"].tolist()
 
    # remove outliers in branch distance 
    outlier_list = outlier_doubleMAD(sub_branch_branch_distance, thresh = 3.5)
    
    indices = [i for i, x in enumerate(outlier_list) if x]
    
    sub_branch_cleaned = sub_branch.drop(sub_branch.index[indices])

    #print(outlier_list)
    #print(indices)
    #print(sub_branch)
    
    print(sub_branch_cleaned)
    
    '''
    min_distance_value_list = sub_branch_cleaned["branch-distance"].tolist()
    
    min_distance_value_list.sort()
    
    min_distance_value = int(min_distance_value_list[2])
    
    print("Smallest branch-distance is:", min_distance_value)
    
    #fig = plt.plot()
    
    (img_endpt_overlay, img_marker) = overlay_skeleton_endpoints(source_image, sub_branch_cleaned)
    
    result_file = (save_path + base_name + '_endpts_overlay' + file_extension)
    plt.savefig(result_file, transparent = True, bbox_inches = 'tight', pad_inches = 0)
    plt.close()
    
    result_file = (save_path + base_name + '_marker' + file_extension)
    cv2.imwrite(result_file, img_marker)
    '''
    
    branch_type_list = sub_branch_cleaned["branch-type"].tolist()
    
    #print(branch_type_list.count(1))
    
    print("[INFO] {} branch end points found\n".format(branch_type_list.count(1)))
    
    #img_hist = branch_data.hist(column = 'branch-distance', by = 'branch-type', bins = 100)
    #result_file = (save_path + base_name + '_hist' + file_extension)
    #plt.savefig(result_file, transparent = True, bbox_inches = 'tight', pad_inches = 0)
    #plt.close()

    
    fig = plt.plot()
    source_image = cv2.cvtColor(orig, cv2.COLOR_BGR2RGB)
    #img_overlay = draw.overlay_euclidean_skeleton_2d(source_image, branch_data, skeleton_color_source = 'branch-distance', skeleton_colormap = 'hsv')
    img_overlay = draw.overlay_euclidean_skeleton_2d(source_image, branch_data, skeleton_color_source = 'branch-type', skeleton_colormap = 'hsv')
    result_file = (save_path + base_name + '_euclidean_graph_overlay' + file_extension)
    plt.savefig(result_file, transparent = True, bbox_inches = 'tight', pad_inches = 0)
    plt.close()
  
     
    ############################################## leaf number computation
    min_distance_value = 20
    #watershed based leaf area segmentaiton 
    labels = watershed_seg(orig, thresh, min_distance_value)
    
    #labels = watershed_seg_marker(orig, thresh, min_distance_value, img_marker)
    
    individual_object_seg(orig, labels, save_path, base_name, file_extension)

    #save watershed result label image
    #Map component labels to hue val
    label_hue = np.uint8(128*labels/np.max(labels))
    #label_hue[labels == largest_label] = np.uint8(15)
    blank_ch = 255*np.ones_like(label_hue)
    labeled_img = cv2.merge([label_hue, blank_ch, blank_ch])

    # cvt to BGR for display
    labeled_img = cv2.cvtColor(labeled_img, cv2.COLOR_HSV2BGR)

    # set background label to black
    labeled_img[label_hue==0] = 0
    result_file = (save_path + base_name + '_label' + file_extension)
    #plt.imsave(result_file, img_as_float(labels), cmap = "Spectral")
    cv2.imwrite(result_file, labeled_img)
    
    (avg_curv, label_trait) = compute_curv(orig, labels)
    
     #save watershed result label image
    result_file = (save_path + base_name + '_curv' + file_extension)
    cv2.imwrite(result_file, label_trait)
    
    
    #find external contour 
    (trait_img, area, solidity, max_width, max_height) = comp_external_contour(image.copy(),thresh)
    # save segmentation result
    result_file = (save_path + base_name + '_excontour' + file_extension)
    #print(filename)
    cv2.imwrite(result_file, trait_img)   
    
    n_leaves = int(len(np.unique(labels))/1 - 1)
    
    #print("[INFO] {} n_leaves found\n".format(len(np.unique(labels)) - 1))
    
    #Path("/tmp/d/a.dat").name
    
    return image_file_name, area, solidity, max_width, max_height, avg_curv, n_leaves


def check_discard_merge(options: List[ImageInput]):
    left = None
    right = None
    i = 0
    replaced = 0
    any_dark = False
    sorted_options = sorted(options, key=lambda o: o.timestamp)
    for option in sorted_options:
        img_name, mean_luminosity, luminosity_str = isbright(option)  # luminosity detection, luminosity_str is either 'dark' or 'bright'
        write_results_to_csv([(img_name, mean_luminosity, luminosity_str)], option.output_directory)
        if luminosity_str == 'dark':
            if left is None:
                left = i
            right = i
            any_dark = True
        elif left is not None and left != right:
            ii = 0
            left_file = sorted_options[left].input_file
            right_file = sorted_options[right].input_file
            prev_file = sorted_options[left - 1].input_file
            next_file = sorted_options[right + 1].input_file
            prev = cv2.imread(prev_file)
            next = cv2.imread(next_file)
            width = right - left + 1
            offset = (1 / width)
            print(f"Replacing {left_file} to {right_file} with merger of {prev_file} and {next_file}")
            for opt in reversed(sorted_options[left:right + 1]):
                prev_weight = ((ii / (right - left)) * ((width - 1) / width)) + offset
                next_weight = ((1 - prev_weight) * ((width - 1) / width)) + offset
                print(f"Merging {prev_file} (weight {round(prev_weight, 2)}) with {next_file} (weight: {round(next_weight, 2)})")
                blended = cv2.addWeighted(prev, prev_weight, next, next_weight, 0)
                cv2.imwrite(opt.input_file, blended)
                cv2.imwrite(join(opt.output_directory, f"{opt.input_stem}.blended.png"), blended)
                ii += 1
            left = None
            right = None
            replaced += width
        i += 1
    print(f"Replaced {replaced} dark images with weighted blends of adjacent images")
    return any_dark


def trait_extract(options: ImageInput) -> ImageResult:
    try:
        _, file_extension = os.path.splitext(options.input_file)
        file_size = os.path.getsize(options.input_file) / MBFACTOR

        print("Segmenting plant object using automatic color clustering method")
        if (file_size > 5.0):
            print(f"It may take some time due to large file size ({file_size} MB)")

        args_colorspace = 'lab'
        args_channels = '1'
        args_num_clusters = 2

        image = cv2.imread(options.input_file)

        # circle detection
        # _, circles, cropped = circle_detect(options)
        image_copy = image.copy()

        # color clustering based plant object segmentation
        segmented = color_cluster_seg(image_copy, args_colorspace, args_channels, args_num_clusters)
        cv2.imwrite(join(options.output_directory, f"{options.input_stem}_seg{file_extension}"), segmented)

        num_clusters = 5
        # save color quantization result
        # rgb_colors = color_quantization(image, thresh, save_path, num_clusters)
        rgb_colors = color_region(image_copy, segmented, options.output_directory + '/', options.input_stem, num_clusters)

        selected_color = rgb2lab(np.uint8(np.asarray([[rgb_colors[0]]])))

        print("Color difference are : ")

        print(selected_color)

        color_diff = []

        for index, value in enumerate(rgb_colors):
            # print(index, value)
            curr_color = rgb2lab(np.uint8(np.asarray([[value]])))
            diff = deltaE_cie76(selected_color, curr_color)

            color_diff.append(diff)

            print(index, value, diff)

            ###############################################

        # accquire medial axis of segmentation mask
        # image_medial_axis = medial_axis_image(thresh)

        image_skeleton, skeleton = skeleton_bw(segmented)

        # save _skeleton result
        cv2.imwrite(join(options.output_directory, f"{options.input_stem}_skeleton{file_extension}"), img_as_ubyte(image_skeleton))

        ###
        # ['skeleton-id', 'node-id-src', 'node-id-dst', 'branch-distance',
        # 'branch-type', 'mean-pixel-value', 'stdev-pixel-value',
        # 'image-coord-src-0', 'image-coord-src-1', 'image-coord-dst-0', 'image-coord-dst-1',
        # 'coord-src-0', 'coord-src-1', 'coord-dst-0', 'coord-dst-1', 'euclidean-distance']
        ###

        # get brach data
        branch_data = summarize(Skeleton(image_skeleton))
        # print(branch_data)

        # select end branch
        sub_branch = branch_data.loc[branch_data['branch-type'] == 1]

        sub_branch_branch_distance = sub_branch["branch-distance"].tolist()

        # remove outliers in branch distance
        outlier_list = outlier_doubleMAD(sub_branch_branch_distance, thresh=3.5)

        indices = [i for i, x in enumerate(outlier_list) if x]

        sub_branch_cleaned = sub_branch.drop(sub_branch.index[indices])

        # print(outlier_list)
        # print(indices)
        # print(sub_branch)

        print(sub_branch_cleaned)

        '''
        min_distance_value_list = sub_branch_cleaned["branch-distance"].tolist()

        min_distance_value_list.sort()

        min_distance_value = int(min_distance_value_list[2])

        print("Smallest branch-distance is:", min_distance_value)

        #fig = plt.plot()

        (img_endpt_overlay, img_marker) = overlay_skeleton_endpoints(source_image, sub_branch_cleaned)

        result_file = (save_path + base_name + '_endpts_overlay' + file_extension)
        plt.savefig(result_file, transparent = True, bbox_inches = 'tight', pad_inches = 0)
        plt.close()

        result_file = (save_path + base_name + '_marker' + file_extension)
        cv2.imwrite(result_file, img_marker)
        '''

        branch_type_list = sub_branch_cleaned["branch-type"].tolist()

        # print(branch_type_list.count(1))

        print("[INFO] {} branch end points found\n".format(branch_type_list.count(1)))

        # img_hist = branch_data.hist(column = 'branch-distance', by = 'branch-type', bins = 100)
        # result_file = (save_path + base_name + '_hist' + file_extension)
        # plt.savefig(result_file, transparent = True, bbox_inches = 'tight', pad_inches = 0)
        # plt.close()

        fig = plt.plot()
        source_image = cv2.cvtColor(image_copy, cv2.COLOR_BGR2RGB)
        # img_overlay = draw.overlay_euclidean_skeleton_2d(source_image, branch_data, skeleton_color_source = 'branch-distance', skeleton_colormap = 'hsv')
        img_overlay = draw.overlay_euclidean_skeleton_2d(source_image, branch_data, skeleton_color_source='branch-type', skeleton_colormap='hsv')
        plt.savefig(join(options.output_directory, f"{options.input_stem}_euclidean_graph_overlay{file_extension}"), transparent=True, bbox_inches='tight', pad_inches=0)
        plt.close()

        ############################################## leaf number computation
        min_distance_value = 20
        # watershed based leaf area segmentaiton
        labels = watershed_seg(image_copy, segmented, min_distance_value)

        # labels = watershed_seg_marker(orig, thresh, min_distance_value, img_marker)

        individual_object_seg(image_copy, labels, options.output_directory + '/', options.input_stem, file_extension)

        # save watershed result label image
        # Map component labels to hue val
        label_hue = np.uint8(128 * labels / np.max(labels))
        # label_hue[labels == largest_label] = np.uint8(15)
        blank_ch = 255 * np.ones_like(label_hue)
        labeled_img = cv2.merge([label_hue, blank_ch, blank_ch])

        # cvt to BGR for display
        labeled_img = cv2.cvtColor(labeled_img, cv2.COLOR_HSV2BGR)

        # set background label to black
        labeled_img[label_hue == 0] = 0
        # plt.imsave(result_file, img_as_float(labels), cmap = "Spectral")
        cv2.imwrite(join(options.output_directory, f"{options.input_stem}_label{file_extension}"), labeled_img)

        (avg_curv, label_trait) = compute_curv(image_copy, labels)

        # save watershed result label image
        cv2.imwrite(join(options.output_directory, f"{options.input_stem}_curv{file_extension}"), label_trait)

        # find external contour
        (trait_img, area, solidity, max_width, max_height) = comp_external_contour(image_copy, segmented)
        # save segmentation result
        # print(filename)
        cv2.imwrite(join(options.output_directory, f"{options.input_stem}_excontour{file_extension}"), trait_img)

        n_leaves = int(len(np.unique(labels)) / 1 - 1)

        # print("[INFO] {} n_leaves found\n".format(len(np.unique(labels)) - 1))

        # Path("/tmp/d/a.dat").name

        return ImageResult(options.input_stem, False, area, solidity, max_width, max_height, avg_curv, n_leaves)
    except:
        print(f"Error in trait extraction: {traceback.format_exc()}")
        return ImageResult(options.input_stem, True, None, None, None, None, None, None)


if __name__ == '__main__':
    
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True,    help="path to image file")
    ap.add_argument("-ft", "--filetype", required=True,    help="Image filetype")
    ap.add_argument("-r", "--result", required = False,    help="result path")
    ap.add_argument('-s', '--color-space', type = str, default ='lab', help='Color space to use: BGR (default), HSV, Lab, YCrCb (YCC)')
    ap.add_argument('-c', '--channels', type = str, default='1', help='Channel indices to use for clustering, where 0 is the first channel,' 
                                                                       + ' 1 is the second channel, etc. E.g., if BGR color space is used, "02" ' 
                                                                       + 'selects channels B and R. (default "all")')
    ap.add_argument('-n', '--num-clusters', type = int, default = 2,  help = 'Number of clusters for K-means clustering (default 2, min 2).')
    args = vars(ap.parse_args())


    # setting path to model file
    file_path = args["path"]
    ext = args['filetype']

    #accquire image file list
    filetype = '*.' + ext
    image_file_path = file_path + filetype
    
    #accquire image file list
    imgList = sorted(glob.glob(image_file_path))

    #print((imgList))
    #global save_path
    
    n_images = len(imgList)
    
    result_list = []
    
    #loop execute
    for image in imgList:
        
        (filename, area, solidity, max_width, max_height, avg_curv, n_leaves) = extract_traits(image)
        
        result_list.append([filename, area, solidity, max_width, max_height, avg_curv, n_leaves])
    
    
    #print(result_list)
    
    
    

    

    '''
    # get cpu number for parallel processing
    agents = psutil.cpu_count()   
    #agents = multiprocessing.cpu_count() 
    #agents = 8
    
    print("Using {0} cores to perfrom parallel processing... \n".format(int(agents)))
    
    # Create a pool of processes. By default, one is created for each CPU in the machine.
    # extract the bouding box for each image in file list
    with closing(Pool(processes = agents)) as pool:
        result_list = pool.map(extract_traits, imgList)
        pool.terminate()
    '''
    
    #trait_file = (os.path.dirname(os.path.abspath(file_path)) + '/' + 'trait.xlsx')
    
    print("Summary: {0} plant images were processed...\n".format(n_images))
    
    #output in command window in a sum table
 
    table = tabulate(result_list, headers = ['filename', 'area', 'solidity', 'max_width', 'max_height' ,'avg_curv', 'n_leaves'], tablefmt = 'orgtbl')

    print(table + "\n")
    
    
    
    
    
    if (args['result']):

        trait_file = (args['result'] + 'trait.xlsx')
        trait_file_csv = (args['result'] + 'trait.csv')
    else:
        trait_file = (file_path + 'trait.xlsx')
        trait_file_csv = (file_path + 'trait.csv')
    
    
    if os.path.isfile(trait_file):
        # update values
        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.active

    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active

        sheet.cell(row = 1, column = 1).value = 'filename'
        sheet.cell(row = 1, column = 2).value = 'leaf_area'
        sheet.cell(row = 1, column = 3).value = 'solidity'
        sheet.cell(row = 1, column = 4).value = 'max_width'
        sheet.cell(row = 1, column = 5).value = 'max_height'
        sheet.cell(row = 1, column = 6).value = 'curvature'
        sheet.cell(row = 1, column = 7).value = 'number_leaf'
    
    for row in result_list:
        sheet.append(row)
   
    #save the csv file
    wb.save(trait_file)
    
    wb = openpyxl.load_workbook(trait_file)
    sh = wb.active # was .get_active_sheet()
    with open(trait_file_csv, 'w', newline = "") as f:
        c = csv.writer(f)
        for r in sh.rows: # generator; was sh.rows
            c.writerow([cell.value for cell in r])
    

    

    
